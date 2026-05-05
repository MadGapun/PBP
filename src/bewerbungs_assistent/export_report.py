"""Bewerbungsbericht — PDF & Excel Export.

Generates comprehensive application reports for:
- Personal tracking and overview
- Arbeitsamt (employment office) documentation
- Self-analysis and optimization

Uses fpdf2 for PDF (pure Python, no system deps).
Uses openpyxl for Excel (optional dependency).
"""

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

from fpdf.enums import XPos, YPos

logger = logging.getLogger("bewerbungs_assistent.export_report")

STATUS_LABELS = {
    "in_vorbereitung": "In Vorbereitung",
    "offen": "Offen",
    "beworben": "Beworben",
    "eingangsbestaetigung": "Eingangsbestätigung",
    "interview": "Interview",
    "zweitgespraech": "Zweitgespräch",
    "interview_abgeschlossen": "Interview abgeschl.",
    "angebot": "Angebot",
    "angenommen": "Angenommen",
    "abgelehnt": "Abgelehnt",
    "zurueckgezogen": "Zurückgezogen",
    "abgelaufen": "Abgelaufen",
}

STATUS_COLORS = {
    "in_vorbereitung": (168, 85, 247),
    "beworben": (56, 189, 248),
    "interview": (251, 191, 36),
    "zweitgespraech": (251, 191, 36),
    "interview_abgeschlossen": (20, 184, 166),
    "angebot": (52, 211, 153),
    "angenommen": (16, 185, 129),
    "abgelehnt": (248, 113, 113),
    "zurueckgezogen": (148, 163, 184),
    "abgelaufen": (148, 163, 184),
    "offen": (100, 116, 139),
    "eingangsbestaetigung": (56, 189, 248),
}

SOURCE_COLORS_PDF = [
    (59, 130, 246), (168, 85, 247), (236, 72, 153), (249, 115, 22),
    (234, 179, 8), (34, 197, 94), (20, 184, 166), (99, 102, 241),
    (244, 63, 94), (107, 114, 128),
]


def _safe_text(text: str) -> str:
    """Sanitize text for PDF output (replace problematic chars)."""
    if not text:
        return ""
    return (text
            .replace("\u2013", "-").replace("\u2014", "-")
            .replace("\u2018", "'").replace("\u2019", "'")
            .replace("\u201c", '"').replace("\u201d", '"')
            .replace("\u2026", "...")
            .replace("\u00ad", "")  # soft hyphen
            .encode("latin-1", errors="replace").decode("latin-1"))


def _line_cell(pdf, w, h, text="", **kwargs):
    """Write a cell and advance to the next line with the modern fpdf2 API."""
    pdf.cell(w, h, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT, **kwargs)


def generate_application_report(report_data: dict, profile: Optional[dict],
                                output_path: Path,
                                zeitraum_von: str = "",
                                zeitraum_bis: str = "",
                                report_settings: Optional[dict] = None) -> Path:
    """Generate a comprehensive PDF Bewerbungsbericht (#173 aufgewertet).

    v1.6.6 (#540): Erweitert um Arbeitsamt-Tauglichkeit. Optionale Bericht-
    Einstellungen (BA-Vermittlungsnummer, Aktenzeichen, Berater-Name,
    Beraterkommentar-Block) werden NUR gerendert, wenn ausgefuellt — der
    Bericht funktioniert weiter fuer Anwender ohne Arbeitsamt-Bezug.
    """
    from fpdf import FPDF

    apps = report_data.get("applications", [])
    stats = report_data.get("statistics", {})
    score_dist = report_data.get("score_distribution", {})
    unapplied = report_data.get("unapplied_high_score", [])
    date_range = report_data.get("date_range", {})
    rejection_patterns = report_data.get("rejection_patterns", {}) or {}
    follow_up_summary = report_data.get("follow_ups", {}) or {}
    bewerbungsart_dist = report_data.get("bewerbungsart", {}) or {}
    # v1.6.5 #540: neue Felder
    aussortiert = report_data.get("stellen_aussortiert", {}) or {}
    interviews_historisch = report_data.get("interviews_historisch_total", 0)
    # v1.6.6 #540: Optionale Settings — Master-Toggle 'arbeitsamt_block_enabled'
    # entscheidet, ob der BA-Block ueberhaupt gerendert wird. So koennen die
    # ausgefuellten Felder im Profil bleiben, der Bericht zeigt sie aber nur
    # wenn der Haken gesetzt ist (User kann den Bericht auch fuer andere
    # Zwecke nutzen, ohne die Felder loeschen zu muessen).
    settings = report_settings or {}
    arbeitsamt_enabled = bool(settings.get("arbeitsamt_block_enabled"))
    ba_vermittlungsnummer = (settings.get("ba_vermittlungsnummer") or "").strip()
    ba_aktenzeichen = (settings.get("ba_aktenzeichen") or "").strip()
    ba_berater_name = (settings.get("ba_berater_name") or "").strip()
    ba_berater_stelle = (settings.get("ba_berater_stelle") or "").strip()
    show_berater_kommentar = bool(settings.get("berater_kommentar_block"))
    # v1.7.0-beta.12 (#582): Taetigkeitsbericht-Modus — Cover-Titel und
    # zusaetzlicher Tagesgrupplerter Aktivitaetsblock; ideal fuer Vermittler-
    # Vorlagen ("Was wurde getan?" statt "Wie laeuft es?").
    taetigkeitsbericht_mode = bool(settings.get("taetigkeitsbericht_mode"))
    # Nur anzeigen wenn Toggle an UND mindestens ein Feld gefuellt
    has_arbeitsamt_block = arbeitsamt_enabled and any([
        ba_vermittlungsnummer, ba_aktenzeichen,
        ba_berater_name, ba_berater_stelle
    ])

    # v1.6.6: Aktivitaetsprotokoll und Quellen-Aktivitaet werden aus den
    # vorhandenen Daten zusammengebaut — kein neuer DB-Aufruf noetig.
    quellen_aktivitaet = report_data.get("scraper_health") or report_data.get("quellen_aktivitaet") or []

    # Zeitraumfilter (#173)
    if zeitraum_von or zeitraum_bis:
        apps = [a for a in apps if
                (not zeitraum_von or (a.get("applied_at") or "") >= zeitraum_von) and
                (not zeitraum_bis or (a.get("applied_at") or "") <= zeitraum_bis + "T23:59:59")]

    # v1.6.6: FPDF-Subclass mit Footer (Erstellt am + Seite X / Y).
    # alias_nb_pages() ersetzt {nb} bei Build-Ende durch die Gesamt-Seitenzahl.
    _erstellt_am_str = (
        f"Erstellt am {datetime.now().strftime('%d.%m.%Y')} um "
        f"{datetime.now().strftime('%H:%M')} Uhr"
    )

    class ReportPDF(FPDF):
        def footer(self):
            # 15mm vom unteren Rand
            self.set_y(-15)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(120, 120, 120)
            self.cell(0, 4, _safe_text(_erstellt_am_str), align="L")
            self.cell(0, 4, _safe_text(f"Seite {self.page_no()} / {{nb}}"),
                     align="R", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            self.set_text_color(0, 0, 0)

    pdf = ReportPDF()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=22)

    # --- Title Page with PBP Branding (#173) ---
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 22)
    cover_title = "Taetigkeitsbericht" if taetigkeitsbericht_mode else "Bewerbungsbericht"
    _line_cell(pdf, 0, 15, _safe_text(cover_title), align="C")

    # Subtitle with name and date range
    pdf.set_font("Helvetica", "", 11)
    name = profile.get("name", "") if profile else ""
    if name:
        _line_cell(pdf, 0, 7, _safe_text(name), align="C")

    # Zeitraum (explicit or derived) + Erstellungsdatum immer prominent zeigen
    pdf.set_font("Helvetica", "B", 10)
    def _de_date(iso: str) -> str:
        if not iso:
            return ""
        try:
            return datetime.fromisoformat(iso[:10]).strftime("%d.%m.%Y")
        except Exception:
            return iso[:10]

    if zeitraum_von or zeitraum_bis:
        zeitraum_label = (
            f"Zeitraum: {_de_date(zeitraum_von) or '–'} bis {_de_date(zeitraum_bis) or 'heute'}"
        )
    elif date_range.get("first") and date_range.get("last"):
        zeitraum_label = (
            f"Zeitraum (aus Daten): {_de_date(date_range['first'])} bis {_de_date(date_range['last'])}"
        )
    else:
        zeitraum_label = "Zeitraum: gesamter Verlauf"
    _line_cell(pdf, 0, 6, _safe_text(zeitraum_label), align="C")
    pdf.set_font("Helvetica", "", 9)
    _line_cell(
        pdf, 0, 5,
        _safe_text(f"Erstellt am {datetime.now().strftime('%d.%m.%Y')} um {datetime.now().strftime('%H:%M')} Uhr"),
        align="C",
    )
    pdf.ln(6)

    # PBP Branding (#173)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(100, 100, 100)
    _line_cell(pdf, 0, 4, _safe_text("Erstellt mit PBP (Persönliches Bewerbungs-Portal)"), align="C")
    _line_cell(pdf, 0, 4, _safe_text("https://github.com/MadGapun/PBP"), align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # v1.6.6 #540: Optionaler Arbeitsamt-Block auf der Cover-Page —
    # nur sichtbar wenn Master-Toggle an UND mindestens ein Feld gefuellt.
    if has_arbeitsamt_block:
        pdf.set_draw_color(180, 180, 180)
        pdf.set_fill_color(245, 247, 250)
        pdf.set_line_width(0.3)
        # Hintergrund-Box mit Padding (left/right margin = 30mm)
        box_start_y = pdf.get_y()
        ba_lines = []
        if ba_vermittlungsnummer:
            ba_lines.append(("Vermittlungsnummer", ba_vermittlungsnummer))
        if ba_aktenzeichen:
            ba_lines.append(("Aktenzeichen", ba_aktenzeichen))
        if ba_berater_name:
            ba_lines.append(("Berater(in)", ba_berater_name))
        if ba_berater_stelle:
            ba_lines.append(("Beratungsstelle", ba_berater_stelle))
        # Box-Hoehe: Header (7) + Zeilen (5 each) + Padding (4)
        box_height = 11 + len(ba_lines) * 5
        pdf.rect(30, box_start_y, 150, box_height, style="DF")
        pdf.set_y(box_start_y + 2)
        pdf.set_font("Helvetica", "B", 10)
        _line_cell(pdf, 0, 6, _safe_text("Vorlage fuer das Arbeitsamt"), align="C")
        for label, value in ba_lines:
            pdf.set_x(36)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(50, 5, _safe_text(f"{label}:"))
            pdf.set_font("Helvetica", "B", 9)
            _line_cell(pdf, 0, 5, _safe_text(value))
        pdf.set_y(box_start_y + box_height + 4)

    # --- Inhaltsverzeichnis (#173) ---
    _section_header(pdf, "Inhaltsverzeichnis")
    toc_items = [
        "1. Executive Summary",
        "2. Bewerbungen nach Status",
        "3. Quellenanalyse",
        "4. Bewerbungsart-Verteilung",
        "5. Fit-Score Verteilung",
        "6. Bewerbungsliste (detailliert)",
        "7. Ablehnungsgruende",
        "8. Offene Follow-ups",
        "9. Nicht beworben trotz gutem Score",
        "10. Keyword-Analyse",
        "11. Aktivitaetsprotokoll",
        "12. Quellen-Aktivitaet (Suchaufwand)",
    ]
    if show_berater_kommentar:
        toc_items.append("13. Beraterkommentar")
    pdf.set_font("Helvetica", "", 10)
    for item in toc_items:
        _line_cell(pdf, 0, 6, _safe_text(f"    {item}"))
    pdf.ln(6)

    # --- Executive Summary (#173) ---
    _section_header(pdf, "1. Executive Summary")
    total_apps = len(apps) if (zeitraum_von or zeitraum_bis) else stats.get("total_applications", len(apps))
    active_jobs = stats.get("active_jobs", 0)
    dismissed_jobs = stats.get("dismissed_jobs", 0)
    avg_score = stats.get("avg_score", 0)
    max_score = stats.get("max_score", 0)

    # Recalculate rates for filtered apps
    # Interview-Zaehler inkl. Folge-Status: wer bei "angebot"/"angenommen" ist,
    # hat zwingend vorher ein Interview gehabt.
    by_status_filtered = Counter(a.get("status", "offen") for a in apps)
    interviews = (by_status_filtered.get("interview", 0)
                  + by_status_filtered.get("zweitgespraech", 0)
                  + by_status_filtered.get("interview_abgeschlossen", 0)
                  + by_status_filtered.get("angebot", 0)
                  + by_status_filtered.get("angenommen", 0))
    offers = by_status_filtered.get("angebot", 0) + by_status_filtered.get("angenommen", 0)
    in_vorb = by_status_filtered.get("in_vorbereitung", 0)
    submitted_apps = total_apps - in_vorb  # exclude in_vorbereitung (#198)
    interview_rate = round(interviews / submitted_apps * 100, 1) if submitted_apps else 0
    offer_rate = round(offers / submitted_apps * 100, 1) if submitted_apps else 0

    # Summary text
    pdf.set_font("Helvetica", "", 9)
    summary_text = (
        f"Im Berichtszeitraum wurden {total_apps} Bewerbungen erfasst. "
        f"Die Interview-Rate liegt bei {interview_rate}%, "
        f"die Angebotsrate bei {offer_rate}%. "
        f"Insgesamt wurden {active_jobs + dismissed_jobs} Stellen analysiert."
    )
    pdf.multi_cell(0, 5, _safe_text(summary_text))
    pdf.ln(2)

    # v1.6.8: „Aktive Filter-Arbeit" und „Geschaetzter Zeitaufwand"
    # entfernt. Begruendung (User-Feedback nach Real-Lauf):
    # - „Aktive Filter-Arbeit" suggeriert „nur 1 Stelle wuerdig befunden",
    #   weil viele Bewerbungen ueber Direct-Add (manuell, ueber Chat)
    #   angelegt werden — nicht ueber stelle_bewerten('passt'). Die Zahl
    #   ist ohne Kontext irrefuehrend.
    # - „Geschaetzter Zeitaufwand" lag mit 63h fuer realen Aufwand um
    #   Groessenordnungen daneben (echter Aufwand: Stunden bis Tage pro
    #   Stelle inkl. Recherche, Anschreiben, Korrekturen, Interview-
    #   Vorbereitung, Dossiers). Heuristik traegt nicht.
    # Beide Bloecke kommen erst zurueck, wenn die Datenbasis stimmt.

    # #430: Monthly application chart in summary
    _embed_chart(pdf, _chart_monthly_bar(apps))

    # --- Pipeline-Kennzahlen (Teil von Executive Summary) ---
    pipeline_data = [
        ("Bewerbungen im Zeitraum", str(total_apps)),
        ("  davon in Vorbereitung", str(in_vorb)),
        ("  davon beworben", str(by_status_filtered.get("beworben", 0))),
        ("  davon im Prozess", str(interviews)),
        ("  davon Angebote", str(offers)),
        ("Aktive Stellen analysiert (gesamt)", str(active_jobs)),
        ("Analysierte Stellen aussortiert (gesamt)", str(dismissed_jobs)),
        ("Durchschn. Fit-Score (alle Jobs)", f"{avg_score}" if avg_score else "k.A."),
        ("Spitzen-Fit-Score (alle Jobs)", f"{max_score}" if max_score else "k.A."),
        ("Interview-Rate (Zeitraum)", f"{interview_rate}%"),
        ("Angebots-Rate (Zeitraum)", f"{offer_rate}%"),
    ]
    for label, value in pipeline_data:
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(80, 5, _safe_text(f"  {label}:"), border=0)
        pdf.set_font("Helvetica", "B", 9)
        _line_cell(pdf, 0, 5, _safe_text(value))
    pdf.ln(4)

    # --- 2. Bewerbungen nach Status ---
    _section_header(pdf, "2. Bewerbungen nach Status")
    by_status = stats.get("applications_by_status", {})
    if by_status:
        # #430: Chart (graceful fallback to colored bars if matplotlib missing)
        _embed_chart(pdf, _chart_status_pie(by_status))
        for status_key, count in sorted(by_status.items(),
                                         key=lambda x: -x[1]):
            label = STATUS_LABELS.get(status_key, status_key)
            bar_width = min(count * 8, 120)
            r, g, b = STATUS_COLORS.get(status_key, (66, 133, 244))
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(45, 5, _safe_text(f"  {label}"))
            pdf.set_fill_color(r, g, b)
            pdf.cell(bar_width, 5, "", fill=True)
            _line_cell(pdf, 15, 5, f"  {count}")
    pdf.ln(4)

    # --- 3. Quellenanalyse (#173) ---
    _section_header(pdf, "3. Quellenanalyse")
    by_source = stats.get("jobs_by_source", {})
    if by_source:
        # #430: Source chart
        _embed_chart(pdf, _chart_source_bar(by_source))
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(60, 5, "  Quelle", border=1, fill=True)
        pdf.cell(30, 5, "Stellen", border=1, fill=True, align="C")
        pdf.cell(40, 5, "Anteil", border=1, fill=True, align="C")
        pdf.ln()
        total_source = sum(by_source.values()) or 1
        pdf.set_font("Helvetica", "", 8)
        for idx, (source, count) in enumerate(sorted(by_source.items(), key=lambda x: -x[1])):
            pct = round(count / total_source * 100, 1)
            r, g, b = SOURCE_COLORS_PDF[idx % len(SOURCE_COLORS_PDF)]
            pdf.set_fill_color(r, g, b)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(5, 5, "", border=0, fill=True)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(55, 5, _safe_text(f" {source}"), border=1)
            pdf.cell(30, 5, str(count), border=1, align="C")
            pdf.cell(40, 5, f"{pct}%", border=1, align="C")
            pdf.ln()
    pdf.ln(2)

    # Bewerbungen nach Quelle mit Erfolgsquote (#173, fix #540 v1.6.5)
    # v1.6.5: has_reached_interview-Flag (#530) statt nur aktueller Status —
    # Bewerbungen die nach Interview auf abgelehnt/abgelaufen rutschten
    # zaehlen jetzt korrekt mit. Vorher zeigte die Tabelle Interviews=0
    # fuer alle Quellen ausser headhunter, obwohl real mehrere durchlaufen
    # waren.
    app_by_source = {}
    for a in apps:
        src = a.get("job_source") or a.get("source") or "unbekannt"
        if src not in app_by_source:
            app_by_source[src] = {"total": 0, "interview": 0}
        app_by_source[src]["total"] += 1
        # Primaere Quelle: has_reached_interview Flag (Schema v29)
        # Fallback: aktueller Status (fuer alte DBs ohne Flag)
        reached_interview = (
            bool(a.get("has_reached_interview"))
            or a.get("status") in ("interview", "zweitgespraech", "interview_abgeschlossen",
                                    "angebot", "angenommen")
        )
        if reached_interview:
            app_by_source[src]["interview"] += 1

    if app_by_source:
        pdf.set_font("Helvetica", "B", 8)
        _line_cell(pdf, 0, 5, _safe_text("  Bewerbungen nach Quelle (mit Erfolgsquote):"))
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(50, 5, "  Quelle", border=1, fill=True)
        pdf.cell(30, 5, "Bewerbungen", border=1, fill=True, align="C")
        pdf.cell(30, 5, "Interviews", border=1, fill=True, align="C")
        pdf.cell(30, 5, "Erfolgsquote", border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for src, data in sorted(app_by_source.items(), key=lambda x: -x[1]["total"]):
            rate = round(data["interview"] / data["total"] * 100, 0) if data["total"] else 0
            pdf.cell(50, 4, _safe_text(f"  {src}"), border=1)
            pdf.cell(30, 4, str(data["total"]), border=1, align="C")
            pdf.cell(30, 4, str(data["interview"]), border=1, align="C")
            pdf.cell(30, 4, f"{rate:.0f}%", border=1, align="C")
            pdf.ln()
    pdf.ln(4)

    # --- 4. Bewerbungsart-Verteilung ---
    if bewerbungsart_dist:
        _section_header(pdf, "4. Bewerbungsart-Verteilung")
        # Filter auf Zeitraum (wenn Zeitraum gesetzt, aus apps neu berechnen)
        art_counter = Counter()
        for a in apps:
            art = (a.get("bewerbungsart") or "unbekannt").strip() or "unbekannt"
            art_counter[art] += 1
        art_total = sum(art_counter.values()) or 1
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(60, 5, "  Art", border=1, fill=True)
        pdf.cell(30, 5, "Anzahl", border=1, fill=True, align="C")
        pdf.cell(30, 5, "Anteil", border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for art, count in sorted(art_counter.items(), key=lambda x: -x[1]):
            pct = round(count / art_total * 100, 1)
            pdf.cell(60, 4, _safe_text(f"  {art}"), border=1)
            pdf.cell(30, 4, str(count), border=1, align="C")
            pdf.cell(30, 4, f"{pct}%", border=1, align="C")
            pdf.ln()
        pdf.ln(4)

    # --- 5. Score-Verteilung ---
    if score_dist:
        _section_header(pdf, "5. Fit-Score Verteilung")
        # #430: Score distribution chart
        _embed_chart(pdf, _chart_score_distribution(score_dist))
        pdf.set_font("Helvetica", "", 9)
        for bracket, count in sorted(score_dist.items()):
            bar_width = min(count, 120)
            pdf.cell(30, 5, _safe_text(f"  Score {bracket}"))
            pdf.set_fill_color(76, 175, 80)
            pdf.cell(bar_width, 5, "", fill=True)
            _line_cell(pdf, 15, 5, f"  {count}")
        pdf.ln(4)

    # --- 6. Bewerbungsliste (detailliert, #173) ---
    _section_header(pdf, "6. Bewerbungsliste (detailliert)")
    if apps:
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(18, 5, "Datum", border=1, fill=True)
        pdf.cell(35, 5, "Firma", border=1, fill=True)
        pdf.cell(45, 5, "Position", border=1, fill=True)
        pdf.cell(22, 5, "Status", border=1, fill=True, align="C")
        pdf.cell(15, 5, "Quelle", border=1, fill=True, align="C")
        pdf.cell(10, 5, "Score", border=1, fill=True, align="C")
        pdf.cell(15, 5, "Art", border=1, fill=True, align="C")
        pdf.cell(30, 5, "Kontakt", border=1, fill=True)
        pdf.ln()
        pdf.set_font("Helvetica", "", 6.5)
        for idx, a in enumerate(apps):
            # Alternating row colors (#173)
            if idx % 2 == 1:
                pdf.set_fill_color(245, 247, 250)
                fill = True
            else:
                fill = False

            date_str = (a.get("applied_at") or "")[:10]
            company = (a.get("company") or "")[:20]
            title = (a.get("title") or "")[:28]
            status_key = a.get("status", "")
            status = STATUS_LABELS.get(status_key, status_key)[:12]
            source = (a.get("job_source") or a.get("source", ""))[:10]
            score = a.get("score", "")
            if a.get("is_pinned"):
                score = f"*{score}" if score else "*"
            else:
                score = str(score) if score else ""
            art = (a.get("bewerbungsart") or "")[:10]
            kontakt = (a.get("ansprechpartner") or "")[:18]

            # Status color badge
            r, g, b = STATUS_COLORS.get(status_key, (100, 116, 139))

            pdf.cell(18, 4, _safe_text(date_str), border=1, fill=fill)
            pdf.cell(35, 4, _safe_text(company), border=1, fill=fill)
            pdf.cell(45, 4, _safe_text(title), border=1, fill=fill)
            # Status with color
            pdf.set_fill_color(r, g, b)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(22, 4, _safe_text(status), border=1, fill=True, align="C")
            pdf.set_text_color(0, 0, 0)
            if idx % 2 == 1:
                pdf.set_fill_color(245, 247, 250)
            pdf.cell(15, 4, _safe_text(source), border=1, fill=fill, align="C")
            pdf.cell(10, 4, _safe_text(str(score)), border=1, fill=fill, align="C")
            pdf.cell(15, 4, _safe_text(art), border=1, fill=fill, align="C")
            pdf.cell(30, 4, _safe_text(kontakt), border=1, fill=fill)
            pdf.ln()
    pdf.ln(4)

    # --- 7. Ablehnungsgruende ---
    _section_header(pdf, "7. Ablehnungsgruende")
    if rejection_patterns and rejection_patterns.get("anzahl"):
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, _safe_text(
            f"  {rejection_patterns['anzahl']} Ablehnung(en) insgesamt."
        ))
        by_reason = rejection_patterns.get("nach_grund", {}) or {}
        if by_reason:
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(255, 228, 230)
            pdf.cell(110, 5, "  Grund", border=1, fill=True)
            pdf.cell(30, 5, "Anzahl", border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 8)
            for grund, count in list(by_reason.items())[:15]:
                grund_txt = (grund or "Kein Grund angegeben")[:70]
                pdf.cell(110, 4, _safe_text(f"  {grund_txt}"), border=1)
                pdf.cell(30, 4, str(count), border=1, align="C")
                pdf.ln()
        ablehnungen = rejection_patterns.get("ablehnungen", []) or []
        if ablehnungen:
            pdf.ln(2)
            pdf.set_font("Helvetica", "B", 8)
            _line_cell(pdf, 0, 5, _safe_text("  Letzte Ablehnungen:"))
            pdf.set_fill_color(255, 228, 230)
            pdf.cell(50, 5, "  Firma", border=1, fill=True)
            pdf.cell(65, 5, "Stelle", border=1, fill=True)
            pdf.cell(25, 5, "Beworben", border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            for r in ablehnungen[:10]:
                pdf.cell(50, 4, _safe_text((r.get("firma") or "")[:28]), border=1)
                pdf.cell(65, 4, _safe_text((r.get("stelle") or "")[:38]), border=1)
                pdf.cell(25, 4, _safe_text((r.get("beworben_am") or "")[:10]), border=1, align="C")
                pdf.ln()
    else:
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, "  Keine Ablehnungen im Zeitraum.")
    pdf.ln(4)

    # --- 7b. Aktiv aussortierte Stellen (#540, v1.6.5) ---
    # User-Wunsch: „die Passt-Nicht-Statistik ist weg" — diese Sub-Sektion
    # zeigt was der User selbst aussortiert hat (Stellen-Filter), als
    # Komplement zu Sektion 7 (Bewerbungs-Ablehnungen durch Recruiter).
    top_dismiss = aussortiert.get("top_gruende", {}) or {}
    if top_dismiss:
        _section_header(pdf, "7b. Eigene Aussortierungen (Stellen)")
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, _safe_text(
            f"  {aussortiert.get('anzahl_total', 0)} Stelle(n) hast du selbst "
            f"als 'passt nicht' markiert — mit folgenden Gruenden:"
        ))
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(255, 243, 224)
        pdf.cell(110, 5, "  Grund", border=1, fill=True)
        pdf.cell(30, 5, "Anzahl", border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for grund, count in list(top_dismiss.items())[:15]:
            grund_txt = (grund or "(kein Grund)")[:70]
            pdf.cell(110, 4, _safe_text(f"  {grund_txt}"), border=1)
            pdf.cell(30, 4, str(count), border=1, align="C")
            pdf.ln()
        pdf.ln(2)
        pdf.set_font("Helvetica", "I", 7)
        pdf.set_text_color(120, 120, 120)
        _line_cell(pdf, 0, 4, _safe_text(
            "  Diese Aussortierungen sind aktive Filter-Arbeit — keine ignorierten Stellen."
        ))
        pdf.set_text_color(0, 0, 0)
        pdf.ln(4)

    # --- 8. Offene Follow-ups ---
    _section_header(pdf, "8. Offene Follow-ups")
    if follow_up_summary and follow_up_summary.get("pending"):
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, _safe_text(
            f"  {follow_up_summary['pending']} offene Follow-ups "
            f"({follow_up_summary.get('overdue', 0)} ueberfaellig)."
        ))
        items = follow_up_summary.get("items", []) or []
        if items:
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(255, 243, 224)
            pdf.cell(50, 5, "  Firma", border=1, fill=True)
            pdf.cell(55, 5, "Stelle", border=1, fill=True)
            pdf.cell(25, 5, "Faellig", border=1, fill=True, align="C")
            pdf.cell(25, 5, "Typ", border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            today_iso = datetime.now().date().isoformat()
            for fu in items:
                sched = (fu.get("scheduled_date") or "")[:10]
                is_overdue = sched and sched < today_iso
                if is_overdue:
                    pdf.set_text_color(200, 60, 60)
                pdf.cell(50, 4, _safe_text((fu.get("firma") or "")[:28]), border=1)
                pdf.cell(55, 4, _safe_text((fu.get("stelle") or "")[:32]), border=1)
                pdf.cell(25, 4, _safe_text(sched), border=1, align="C")
                pdf.cell(25, 4, _safe_text((fu.get("follow_up_type") or "")[:12]), border=1, align="C")
                pdf.ln()
                if is_overdue:
                    pdf.set_text_color(0, 0, 0)
    else:
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, "  Keine offenen Follow-ups.")
    pdf.ln(4)

    # --- 9. Nicht beworben trotz gutem Score (#220, fix #532 v1.6.4) ---
    _section_header(pdf, "9. Nicht beworben trotz gutem Fit-Score")
    if unapplied:
        pdf.set_font("Helvetica", "", 8)
        # v1.6.4: Header zeigt was wirklich in der Tabelle landet, nicht Pool-Groesse
        max_rows = 30
        shown = min(len(unapplied), max_rows)
        if len(unapplied) > max_rows:
            header_text = (
                f"  Top {max_rows} von {len(unapplied)} aktiven Stellen mit Score >= 5 "
                f"ohne Bewerbung (sortiert nach Score):"
            )
        else:
            header_text = (
                f"  {len(unapplied)} aktive Stelle{'n' if len(unapplied) != 1 else ''} "
                f"mit Score >= 5 ohne Bewerbung:"
            )
        _line_cell(pdf, 0, 5, _safe_text(header_text))
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(255, 243, 224)
        pdf.cell(55, 5, "Firma", border=1, fill=True)
        pdf.cell(70, 5, "Position", border=1, fill=True)
        pdf.cell(13, 5, "Score", border=1, fill=True, align="C")
        pdf.cell(22, 5, "Quelle", border=1, fill=True, align="C")
        pdf.cell(22, 5, "Gefunden", border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 7)
        # v1.6.4: cap to max_rows, keine Grund-Spalte mehr (DB liefert nur is_active=1)
        for j in unapplied[:max_rows]:
            company = (j.get("company") or "")[:32]
            title = (j.get("title") or "")[:42]
            pdf.cell(55, 4, _safe_text(company), border=1)
            pdf.cell(70, 4, _safe_text(title), border=1)
            pdf.cell(13, 4, str(j.get("score", "")), border=1, align="C")
            pdf.cell(22, 4, _safe_text((j.get("source") or "")[:14]), border=1, align="C")
            pdf.cell(22, 4, _safe_text((j.get("found_at") or "")[:10]), border=1, align="C")
            pdf.ln()
    else:
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, "  Keine offenen Stellen mit gutem Score (alle bewertet oder beworben).")
    pdf.ln(4)

    # --- 10. Keyword-Analyse ---
    _section_header(pdf, "10. Keyword-Analyse (Top-Begriffe in passenden Stellen)")
    applied_descriptions = []
    for a in apps:
        desc = a.get("job_description") or ""
        title = a.get("title") or ""
        if desc or title:
            applied_descriptions.append(f"{title} {desc}".lower())

    if applied_descriptions:
        # Count word frequency across applied job descriptions
        word_counter = Counter()
        stop_words = {
            "und", "die", "der", "in", "von", "zu", "mit", "für", "für",
            "den", "das", "ist", "im", "ein", "eine", "auf", "des", "als",
            "wir", "sie", "oder", "an", "bei", "our", "the", "and", "for",
            "you", "are", "with", "your", "will", "that", "this", "have",
            "from", "not", "all", "can", "has", "been", "more", "also",
            "aber", "aus", "nach", "wie", "sich", "ihre", "ihren", "einer",
            "einem", "eines", "werden", "wird", "haben", "sein", "sind",
            "m/w/d", "m/w", "(m/w/d)", "(m/w)", "gmbh", "gerne",
        }
        for text in applied_descriptions:
            words = text.split()
            for w in words:
                w = w.strip(".,;:!?()[]{}\"'/-")
                if len(w) >= 3 and w not in stop_words:
                    word_counter[w] += 1

        top_words = word_counter.most_common(25)
        if top_words:
            pdf.set_font("Helvetica", "", 8)
            _line_cell(pdf, 0, 5, _safe_text(
                "  Häufigste Begriffe in Stellen, auf die Sie sich beworben haben:"
            ))
            pdf.set_font("Helvetica", "B", 7)
            pdf.set_fill_color(232, 245, 233)
            pdf.cell(50, 5, "  Keyword", border=1, fill=True)
            pdf.cell(25, 5, "Vorkommen", border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 7)
            for word, count in top_words:
                pdf.cell(50, 4, _safe_text(f"  {word}"), border=1)
                pdf.cell(25, 4, str(count), border=1, align="C")
                pdf.ln()
    pdf.ln(4)

    # === v1.6.6 #540: Neue Sektionen 11, 12, 13 (+ optional 14) ===

    # --- 11. Aktivitaetsprotokoll (chronologische Timeline) ---
    pdf.add_page()
    _section_header(pdf, "11. Aktivitaetsprotokoll")
    pdf.set_font("Helvetica", "", 8)
    _line_cell(pdf, 0, 5, _safe_text(
        "  Chronologie aller wichtigen Bewerbungs-Ereignisse im Berichtszeitraum."
    ))
    pdf.ln(1)
    timeline_events = []
    for a in apps:
        applied_at = (a.get("applied_at") or "")[:10]
        title = a.get("title") or ""
        company = a.get("company") or ""
        status = STATUS_LABELS.get(a.get("status") or "", a.get("status") or "")
        if applied_at:
            timeline_events.append((
                applied_at, "Bewerbung", f"{company} - {title}", status,
            ))
        last_event_at = (a.get("last_event_at") or a.get("updated_at") or "")[:10]
        if last_event_at and last_event_at != applied_at and a.get("status"):
            timeline_events.append((
                last_event_at, "Statuswechsel", f"{company} - {title}", status,
            ))
    timeline_events.sort(reverse=True)
    if timeline_events:
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(22, 5, "Datum", border=1, fill=True, align="C")
        pdf.cell(28, 5, "Ereignis", border=1, fill=True, align="C")
        pdf.cell(95, 5, "Bewerbung", border=1, fill=True)
        pdf.cell(35, 5, "Status", border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 7)
        for date, evt, target, status in timeline_events[:60]:
            try:
                date_de = datetime.fromisoformat(date).strftime("%d.%m.%Y")
            except Exception:
                date_de = date
            pdf.cell(22, 4, date_de, border=1, align="C")
            pdf.cell(28, 4, _safe_text(evt[:18]), border=1, align="C")
            pdf.cell(95, 4, _safe_text(target[:60]), border=1)
            pdf.cell(35, 4, _safe_text(status[:22]), border=1, align="C")
            pdf.ln()
        if len(timeline_events) > 60:
            pdf.set_font("Helvetica", "I", 7)
            _line_cell(pdf, 0, 4, _safe_text(
                f"  ... weitere {len(timeline_events) - 60} Ereignisse nicht angezeigt."
            ))
    else:
        _line_cell(pdf, 0, 5, _safe_text(
            "  Keine Aktivitaeten im Berichtszeitraum erfasst."
        ))
    pdf.ln(4)

    # v1.7.0-beta.12 (#582): Tagesgruppierte Aktivitaets-Uebersicht —
    # nur im Taetigkeitsbericht-Modus. Buendelt alle Ereignisse pro Tag,
    # so dass Vermittler/Berater auf einen Blick sehen, an welchen Tagen
    # aktiv gearbeitet wurde und was konkret passiert ist.
    if taetigkeitsbericht_mode and timeline_events:
        from collections import defaultdict as _dd
        per_day = _dd(list)
        for date, evt, target, status in timeline_events:
            per_day[date].append((evt, target, status))
        pdf.add_page()
        _section_header(pdf, "11a. Taegliche Aktivitaets-Uebersicht")
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, _safe_text(
            "  Aktivitaeten gebuendelt pro Tag — als Nachweis konkreter Bemuehungen."
        ))
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(28, 5, "Datum", border=1, fill=True, align="C")
        pdf.cell(15, 5, "Aktionen", border=1, fill=True, align="C")
        pdf.cell(137, 5, "Details", border=1, fill=True)
        pdf.ln()
        pdf.set_font("Helvetica", "", 7)
        for day in sorted(per_day.keys(), reverse=True)[:90]:
            try:
                day_de = datetime.fromisoformat(day).strftime("%d.%m.%Y")
            except Exception:
                day_de = day
            entries = per_day[day]
            details = "; ".join(
                f"{evt}: {target}" for evt, target, _status in entries[:3]
            )
            if len(entries) > 3:
                details += f"; +{len(entries) - 3} weitere"
            pdf.cell(28, 4, day_de, border=1, align="C")
            pdf.cell(15, 4, str(len(entries)), border=1, align="C")
            pdf.cell(137, 4, _safe_text(details[:90]), border=1)
            pdf.ln()
        pdf.ln(4)

    # --- 12. Quellen-Aktivitaet (Suchaufwand) ---
    _section_header(pdf, "12. Quellen-Aktivitaet (Suchaufwand)")
    pdf.set_font("Helvetica", "", 8)
    _line_cell(pdf, 0, 5, _safe_text(
        "  Belegt den Suchaufwand: welche Job-Portale wurden durchsucht, "
        "wie haeufig, mit welcher Trefferzahl."
    ))
    pdf.ln(1)
    if quellen_aktivitaet:
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(230, 230, 230)
        pdf.cell(40, 5, "Quelle", border=1, fill=True)
        pdf.cell(25, 5, "Suchlaeufe", border=1, fill=True, align="C")
        pdf.cell(25, 5, "Erfolgreich", border=1, fill=True, align="C")
        pdf.cell(28, 5, "Letzte Treffer", border=1, fill=True, align="C")
        pdf.cell(30, 5, "Letzter Lauf", border=1, fill=True, align="C")
        pdf.cell(20, 5, "Status", border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 7)
        for q in quellen_aktivitaet:
            name = (q.get("scraper_name") or q.get("name") or "")[:18]
            total_runs = q.get("total_runs") or 0
            total_succ = q.get("total_successes") or 0
            last_count = q.get("last_count") or q.get("letzte_rohtreffer") or 0
            last_run = (q.get("last_run") or "")[:10]
            try:
                last_run_de = datetime.fromisoformat(last_run).strftime("%d.%m.%Y") if last_run else "-"
            except Exception:
                last_run_de = last_run or "-"
            aktiv = "aktiv" if q.get("is_active") else "inaktiv"
            pdf.cell(40, 4, _safe_text(name), border=1)
            pdf.cell(25, 4, str(total_runs), border=1, align="C")
            pdf.cell(25, 4, str(total_succ), border=1, align="C")
            pdf.cell(28, 4, str(last_count), border=1, align="C")
            pdf.cell(30, 4, last_run_de, border=1, align="C")
            pdf.cell(20, 4, aktiv, border=1, align="C")
            pdf.ln()
    else:
        _line_cell(pdf, 0, 5, _safe_text(
            "  Keine Quellen-Aktivitaet erfasst (noch keine Suchen durchgefuehrt)."
        ))
    pdf.ln(4)

    # v1.6.8: Sektion 13 „Bewerbungs-Trichter" entfernt — die Stufen waren
    # in sich nicht schluessig (1027 aussortiert + 68 beworben passte nicht
    # zu 1028 gesichtet, weil Bewerbungen auch ueber Direct-Add aus
    # externen Quellen kommen, nicht nur aus dem gesichteten Pool). Kommt
    # zurueck wenn die Datenmodellierung das sauber abbildet.

    # --- 13. Beraterkommentar (optional) ---
    if show_berater_kommentar:
        _section_header(pdf, "13. Beraterkommentar")
        pdf.set_font("Helvetica", "", 8)
        _line_cell(pdf, 0, 5, _safe_text(
            "  Platz fuer handschriftliche Anmerkungen oder gemeinsame Notizen "
            "aus dem Beratungsgespraech."
        ))
        pdf.ln(2)
        pdf.set_draw_color(180, 180, 180)
        pdf.set_line_width(0.2)
        # 8 leere Zeilen mit Trennlinien
        line_y = pdf.get_y() + 4
        for _ in range(8):
            pdf.line(15, line_y, 195, line_y)
            line_y += 7
        pdf.set_y(line_y)
        pdf.ln(2)

    # v1.6.6: Old closing-footer-block entfernt — Cover-Page-Branding +
    # Per-Seite-Footer (Erstellt am + Seite X/Y) reichen aus.
    pdf.output(str(output_path))
    logger.info("PDF Bewerbungsbericht erstellt: %s", output_path)
    return output_path


def generate_data_self_disclosure(db, profile: Optional[dict],
                                    output_path: Path) -> Path:
    """Erstellt eine PDF-Datenauskunft fuer den User (v1.7.0 #581).

    Listet ALLE Daten die PBP ueber den User gespeichert hat — ohne
    sensitive Inhalte. Geeignet fuer DSGVO Art. 15 / persoenliche
    Selbstauskunft.
    """
    from fpdf import FPDF
    from .database import get_data_dir
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 18)
    _line_cell(pdf, 0, 12, _safe_text("Datenauskunft (DSGVO Art. 15)"), align="C")
    pdf.set_font("Helvetica", "", 10)
    name = profile.get("name") if profile else "(kein Profil)"
    _line_cell(pdf, 0, 6, _safe_text(name or "(ohne Namen)"), align="C")
    _line_cell(
        pdf, 0, 5,
        _safe_text(f"Erstellt am {datetime.now().strftime('%d.%m.%Y')}"),
        align="C",
    )
    pdf.ln(8)

    # Section: Persoenliche Daten
    _section_header(pdf, "1. Persoenliche Daten")
    pdf.set_font("Helvetica", "", 9)
    if profile:
        fields = [
            ("Name", profile.get("name")),
            ("E-Mail", profile.get("email")),
            ("Telefon", profile.get("phone")),
            ("Adresse", profile.get("address")),
            ("Stadt", profile.get("city")),
            ("PLZ", profile.get("plz")),
            ("Land", profile.get("country")),
            ("Geburtsdatum", profile.get("birthday")),
            ("Nationalitaet", profile.get("nationality")),
        ]
        for label, val in fields:
            if val:
                pdf.cell(50, 5, _safe_text(f"  {label}:"))
                _line_cell(pdf, 0, 5, _safe_text(str(val)))
    else:
        _line_cell(pdf, 0, 5, "  Kein Profil vorhanden.")
    pdf.ln(4)

    # Counts (no contents)
    _section_header(pdf, "2. Datenumfang (nur Anzahlen, keine Inhalte)")
    pdf.set_font("Helvetica", "", 9)
    counts = []
    try:
        conn = db.connect()
        pid = db.get_active_profile_id()
        for label, sql in [
            ("Bewerbungen",
             "SELECT COUNT(*) FROM applications WHERE profile_id=? OR profile_id IS NULL"),
            ("Stellen",
             "SELECT COUNT(*) FROM jobs WHERE profile_id=? OR profile_id IS NULL"),
            ("Dokumente (Metadaten)",
             "SELECT COUNT(*) FROM documents WHERE profile_id=? OR profile_id IS NULL"),
            ("Termine/Meetings",
             "SELECT COUNT(*) FROM application_meetings"),
            ("E-Mails",
             "SELECT COUNT(*) FROM application_emails"),
            ("Skills",
             "SELECT COUNT(*) FROM skills WHERE profile_id=?"),
            ("Kontakte",
             "SELECT COUNT(*) FROM contacts WHERE profile_id=? OR profile_id IS NULL"),
            ("Follow-ups",
             "SELECT COUNT(*) FROM follow_ups"),
        ]:
            try:
                params = (pid, pid) if "?" in sql and sql.count("?") == 2 else ((pid,) if "?" in sql else ())
                row = conn.execute(sql, params).fetchone()
                n = row[0] if row else 0
                counts.append((label, n))
            except Exception:
                counts.append((label, "—"))
    except Exception:
        pass
    for label, n in counts:
        pdf.cell(60, 5, _safe_text(f"  {label}:"))
        _line_cell(pdf, 0, 5, str(n))
    pdf.ln(4)

    # Speicher-Orte
    _section_header(pdf, "3. Speicher-Orte")
    pdf.set_font("Helvetica", "", 9)
    try:
        data_dir = get_data_dir()
        _line_cell(pdf, 0, 5, _safe_text(f"  Datenbank: {data_dir / 'pbp.db'}"))
        _line_cell(pdf, 0, 5, _safe_text(f"  Dokumente: {data_dir / 'dokumente'}"))
        _line_cell(pdf, 0, 5, _safe_text(f"  Backups:   {data_dir / 'backups'}"))
    except Exception as e:
        _line_cell(pdf, 0, 5, f"  Pfad konnte nicht ermittelt werden: {e}")
    pdf.ln(4)

    # Externalisierung
    _section_header(pdf, "4. Daten-Externalisierung")
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 4.5, _safe_text(
        "  PBP speichert Daten standardmaessig nur lokal auf deinem Geraet. "
        "Externe Dienste werden nur dann genutzt, wenn du sie explizit aktivierst:"
    ))
    pdf.ln(1)
    _line_cell(pdf, 0, 4.5, _safe_text(
        "  - Aktive Job-Quellen: senden anonymisierte Suchanfragen "
        "(Keywords, Region) an die jeweiligen Job-Portale."
    ))
    _line_cell(pdf, 0, 4.5, _safe_text(
        "  - Claude (MCP): nur Daten, die du Claude aktiv im Chat zur "
        "Verfuegung stellst."
    ))
    _line_cell(pdf, 0, 4.5, _safe_text(
        "  - Lokale AI (Ollama, optional): laeuft komplett auf deinem Geraet, "
        "keine Daten verlassen den PC."
    ))
    pdf.ln(4)

    # Datenschutz-Hinweis
    _section_header(pdf, "5. Hinweise")
    pdf.set_font("Helvetica", "", 8)
    pdf.multi_cell(0, 4, _safe_text(
        "  Diese PDF enthaelt keine Passwoerter, API-Keys oder vollstaendige "
        "Dokumenten-Inhalte — nur Metadaten und Anzahlen. Sie ist gedacht zur "
        "eigenen Verwendung (Beratungsgespraech, Datenschutz-Behoerde, oder "
        "zur Erklaerung an Familie/Freunde was PBP ueber dich speichert)."
    ))
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(120, 120, 120)
    _line_cell(pdf, 0, 4, _safe_text(
        f"  Erstellt mit PBP — github.com/MadGapun/PBP — {datetime.now().isoformat(timespec='minutes')}"
    ))

    pdf.output(str(output_path))
    logger.info("PDF Datenauskunft erstellt: %s", output_path)
    return output_path


def _section_header(pdf, title: str):
    """Draw a section header with background."""
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_fill_color(33, 150, 243)
    pdf.set_text_color(255, 255, 255)
    _line_cell(pdf, 0, 7, _safe_text(f"  {title}"), fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)


# --- Chart helpers (#430) ---

def _has_matplotlib() -> bool:
    """Check if matplotlib is available."""
    try:
        import matplotlib
        return True
    except ImportError:
        return False


def _chart_to_bytes(fig) -> bytes:
    """Render a matplotlib figure to PNG bytes."""
    from io import BytesIO
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=fig.get_facecolor(), edgecolor="none")
    buf.seek(0)
    return buf.read()


def _chart_status_pie(by_status: dict) -> bytes | None:
    """Pie chart of application status distribution."""
    if not by_status or not _has_matplotlib():
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        labels = [STATUS_LABELS.get(k, k) for k in by_status]
        sizes = list(by_status.values())
        colors = [
            "#{:02x}{:02x}{:02x}".format(*STATUS_COLORS.get(k, (100, 116, 139)))
            for k in by_status
        ]

        fig, ax = plt.subplots(figsize=(4.5, 3))
        fig.patch.set_facecolor("white")
        wedges, texts, autotexts = ax.pie(
            sizes, labels=labels, colors=colors, autopct="%1.0f%%",
            startangle=90, textprops={"fontsize": 7},
        )
        for t in autotexts:
            t.set_fontsize(7)
            t.set_color("white")
            t.set_fontweight("bold")
        ax.set_title("Bewerbungen nach Status", fontsize=9, fontweight="bold")
        data = _chart_to_bytes(fig)
        plt.close(fig)
        return data
    except Exception as e:
        logger.debug("Chart status_pie fehlgeschlagen: %s", e)
        return None


def _chart_monthly_bar(apps: list) -> bytes | None:
    """Bar chart of applications per month."""
    if not apps or not _has_matplotlib():
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        months = Counter()
        for a in apps:
            dt = (a.get("applied_at") or "")[:7]  # YYYY-MM
            if dt and len(dt) == 7:
                months[dt] += 1
        if not months:
            return None

        sorted_months = sorted(months.keys())
        labels = sorted_months
        values = [months[m] for m in sorted_months]

        fig, ax = plt.subplots(figsize=(5, 2.8))
        fig.patch.set_facecolor("white")
        bars = ax.bar(labels, values, color="#2196F3", width=0.6)
        ax.set_title("Bewerbungen pro Monat", fontsize=9, fontweight="bold")
        ax.set_ylabel("Anzahl", fontsize=8)
        ax.tick_params(axis="x", rotation=45, labelsize=7)
        ax.tick_params(axis="y", labelsize=7)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                    str(val), ha="center", va="bottom", fontsize=7)
        fig.tight_layout()
        data = _chart_to_bytes(fig)
        plt.close(fig)
        return data
    except Exception as e:
        logger.debug("Chart monthly_bar fehlgeschlagen: %s", e)
        return None


def _chart_source_bar(by_source: dict) -> bytes | None:
    """Horizontal bar chart of jobs by source."""
    if not by_source or not _has_matplotlib():
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        sorted_items = sorted(by_source.items(), key=lambda x: x[1])
        labels = [s[0] for s in sorted_items]
        values = [s[1] for s in sorted_items]
        colors = [
            "#{:02x}{:02x}{:02x}".format(*SOURCE_COLORS_PDF[i % len(SOURCE_COLORS_PDF)])
            for i in range(len(labels))
        ]

        fig, ax = plt.subplots(figsize=(4.5, max(2, len(labels) * 0.4 + 1)))
        fig.patch.set_facecolor("white")
        bars = ax.barh(labels, values, color=colors, height=0.6)
        ax.set_title("Stellen nach Quelle", fontsize=9, fontweight="bold")
        ax.set_xlabel("Anzahl", fontsize=8)
        ax.tick_params(axis="both", labelsize=7)
        for bar, val in zip(bars, values):
            ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2,
                    str(val), ha="left", va="center", fontsize=7)
        fig.tight_layout()
        data = _chart_to_bytes(fig)
        plt.close(fig)
        return data
    except Exception as e:
        logger.debug("Chart source_bar fehlgeschlagen: %s", e)
        return None


def _chart_score_distribution(score_dist: dict) -> bytes | None:
    """Bar chart of fit-score distribution."""
    if not score_dist or not _has_matplotlib():
        return None
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        sorted_brackets = sorted(score_dist.keys())
        labels = [f"Score {b}" for b in sorted_brackets]
        values = [score_dist[b] for b in sorted_brackets]

        fig, ax = plt.subplots(figsize=(4.5, 2.5))
        fig.patch.set_facecolor("white")
        bars = ax.bar(labels, values, color="#4CAF50", width=0.6)
        ax.set_title("Fit-Score Verteilung", fontsize=9, fontweight="bold")
        ax.set_ylabel("Anzahl Stellen", fontsize=8)
        ax.tick_params(axis="both", labelsize=7)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                    str(val), ha="center", va="bottom", fontsize=7)
        fig.tight_layout()
        data = _chart_to_bytes(fig)
        plt.close(fig)
        return data
    except Exception as e:
        logger.debug("Chart score_distribution fehlgeschlagen: %s", e)
        return None


def _embed_chart(pdf, chart_data: bytes | None, max_width: int = 170):
    """Embed a PNG chart into the PDF if data is available."""
    if not chart_data:
        return
    import tempfile
    import os
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
        f.write(chart_data)
        tmp_path = f.name
    try:
        # Check if chart would overflow page
        if pdf.get_y() > 220:
            pdf.add_page()
        pdf.image(tmp_path, x=20, w=max_width)
        pdf.ln(4)
    except Exception as e:
        logger.debug("Chart-Einbettung fehlgeschlagen: %s", e)
    finally:
        os.unlink(tmp_path)


def generate_excel_report(report_data: dict, profile: Optional[dict],
                          output_path: Path,
                          zeitraum_von: str = "",
                          zeitraum_bis: str = "",
                          report_settings: Optional[dict] = None) -> Path:
    """Generate an Excel Bewerbungsbericht (requires openpyxl).

    v1.6.6 (#540): Akzeptiert report_settings fuer Signatur-Parity mit dem
    PDF-Generator. Aktuell wird `report_settings` im Excel-Export nicht
    visuell genutzt (Excel ist primaer fuer Daten-Auswertung gedacht), die
    Signatur ist aber bereit fuer kuenftige Erweiterung.
    """
    _ = report_settings  # vorgehalten fuer kuenftige Excel-Erweiterungen
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    apps = report_data.get("applications", [])
    stats = report_data.get("statistics", {})

    # Zeitraumfilter (Parity mit PDF)
    if zeitraum_von or zeitraum_bis:
        apps = [a for a in apps if
                (not zeitraum_von or (a.get("applied_at") or "") >= zeitraum_von) and
                (not zeitraum_bis or (a.get("applied_at") or "") <= zeitraum_bis + "T23:59:59")]

    wb = Workbook()

    # --- Sheet 1: Bewerbungen ---
    ws = wb.active
    ws.title = "Bewerbungen"
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["Datum", "Firma", "Position", "Status", "Quelle",
               "Fit-Score", "Gepinnt", "Bewerbungsart", "Notizen"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_num, a in enumerate(apps, 2):
        values = [
            (a.get("applied_at") or "")[:10],
            a.get("company", ""),
            a.get("title", ""),
            STATUS_LABELS.get(a.get("status", ""), a.get("status", "")),
            a.get("job_source") or "",
            a.get("score", 0) if not a.get("is_pinned") else "",
            "Ja" if a.get("is_pinned") else "",
            a.get("bewerbungsart", ""),
            (a.get("notes") or "")[:200],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # --- Sheet 2: Statistik ---
    ws2 = wb.create_sheet("Statistik")

    # Zeitraum + Erstellt am (kopfzeile)
    def _de_date_xlsx(iso: str) -> str:
        if not iso:
            return ""
        try:
            return datetime.fromisoformat(iso[:10]).strftime("%d.%m.%Y")
        except Exception:
            return iso[:10]

    meta_rows = []
    if zeitraum_von or zeitraum_bis:
        meta_rows.append(("Zeitraum",
                          f"{_de_date_xlsx(zeitraum_von) or '-'} bis "
                          f"{_de_date_xlsx(zeitraum_bis) or 'heute'}"))
    date_range = report_data.get("date_range") or {}
    if not meta_rows and date_range.get("first") and date_range.get("last"):
        meta_rows.append(("Zeitraum (aus Daten)",
                          f"{_de_date_xlsx(date_range['first'])} bis "
                          f"{_de_date_xlsx(date_range['last'])}"))
    meta_rows.append(("Erstellt am",
                      datetime.now().strftime("%d.%m.%Y %H:%M")))
    for idx, (label, val) in enumerate(meta_rows, 1):
        c = ws2.cell(row=idx, column=1, value=label)
        c.font = Font(bold=True)
        ws2.cell(row=idx, column=2, value=val)

    meta_offset = len(meta_rows) + 1  # Leerzeile zwischen Meta und Stats

    stat_rows = [
        ("Bewerbungen gesamt", stats.get("total_applications", 0)),
        ("Aktive Stellen", stats.get("active_jobs", 0)),
        ("Analysierte Stellen (aussortiert)", stats.get("dismissed_jobs", 0)),
        ("Durchschn. Fit-Score", stats.get("avg_score", "")),
        ("Spitzen-Score", stats.get("max_score", "")),
        ("Interview-Rate", f"{stats.get('interview_rate', 0)}%"),
        ("Angebots-Rate", f"{stats.get('offer_rate', 0)}%"),
    ]
    header_row = meta_offset + 1
    ws2.cell(row=header_row, column=1, value="Kennzahl").font = Font(bold=True)
    ws2.cell(row=header_row, column=2, value="Wert").font = Font(bold=True)
    for row_num, (label, val) in enumerate(stat_rows, header_row + 1):
        ws2.cell(row=row_num, column=1, value=label)
        ws2.cell(row=row_num, column=2, value=val)

    # Status breakdown
    row_start = header_row + len(stat_rows) + 2
    ws2.cell(row=row_start, column=1, value="Status-Verteilung").font = Font(bold=True)
    for i, (status, count) in enumerate(
        stats.get("applications_by_status", {}).items(), row_start + 1
    ):
        ws2.cell(row=i, column=1, value=STATUS_LABELS.get(status, status))
        ws2.cell(row=i, column=2, value=count)

    # Sources
    source_start = row_start + len(stats.get("applications_by_status", {})) + 2
    ws2.cell(row=source_start, column=1, value="Jobquellen").font = Font(bold=True)
    for i, (source, count) in enumerate(
        stats.get("jobs_by_source", {}).items(), source_start + 1
    ):
        ws2.cell(row=i, column=1, value=source)
        ws2.cell(row=i, column=2, value=count)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    # #430: Excel charts (openpyxl built-in)
    try:
        from openpyxl.chart import PieChart, BarChart, Reference

        by_status = stats.get("applications_by_status", {})
        if by_status and row_start > 3:
            pie = PieChart()
            pie.title = "Bewerbungen nach Status"
            pie.style = 10
            status_count = len(by_status)
            cats = Reference(ws2, min_col=1, min_row=row_start + 1,
                             max_row=row_start + status_count)
            vals = Reference(ws2, min_col=2, min_row=row_start + 1,
                             max_row=row_start + status_count)
            pie.add_data(vals, titles_from_data=False)
            pie.set_categories(cats)
            pie.width = 14
            pie.height = 10
            ws2.add_chart(pie, "D2")

        by_source = stats.get("jobs_by_source", {})
        if by_source:
            bar = BarChart()
            bar.type = "col"
            bar.title = "Stellen nach Quelle"
            bar.style = 10
            bar.y_axis.title = "Anzahl"
            source_count = len(by_source)
            cats = Reference(ws2, min_col=1, min_row=source_start + 1,
                             max_row=source_start + source_count)
            vals = Reference(ws2, min_col=2, min_row=source_start + 1,
                             max_row=source_start + source_count)
            bar.add_data(vals, titles_from_data=False)
            bar.set_categories(cats)
            bar.width = 14
            bar.height = 10
            ws2.add_chart(bar, "D16")
    except Exception as e:
        logger.debug("Excel-Charts fehlgeschlagen: %s", e)

    wb.save(str(output_path))
    logger.info("Excel Bewerbungsbericht erstellt: %s", output_path)
    return output_path
